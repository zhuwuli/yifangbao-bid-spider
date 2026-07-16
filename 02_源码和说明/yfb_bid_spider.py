from __future__ import annotations

import argparse
import base64
import copy
import json
import os
import re
import shutil
import sys
import threading
import time
import uuid
from concurrent.futures import ThreadPoolExecutor, as_completed
from datetime import datetime, timedelta
from pathlib import Path
from typing import Any
from urllib.parse import urlencode
from urllib.request import Request, urlopen

import requests
from openpyxl import load_workbook
from openpyxl.styles import Font
from openpyxl.worksheet.hyperlink import Hyperlink
from requests.adapters import HTTPAdapter


BASE = "https://qiye.qianlima.com/new_qd_yfbsite/api"
REFERER = "https://qiye.qianlima.com/new_qd_yfbsite/#/infoCenter/search"
AREA_IDS = "1738,1740"  # 济南, 莱芜
KEYWORD = "监测"
SEARCH_KEYWORDS = ("监测", "水土保持", "测绘", "测量", "绿色建筑评价", "绿色建筑验收")
TITLE_KEEP_TERMS = (
    "水土保持", "水保", "绿色建筑评价", "绿色建筑验收", "绿建评价", "绿建验收", "绿色建筑",
    "测绘", "测量", "多测合一", "地形图", "规划核实", "房产实测", "国土变更调查",
    "不动产测绘", "竣工测量", "放线测量", "验线测量",
    "基坑监测", "深基坑监测", "第三方监测", "专项监测", "变形监测", "沉降观测", "沉降监测",
    "结构监测", "运营期结构监测", "监测服务",
)
WATER_GREEN_TERMS = ("水土保持", "水保", "绿色建筑评价", "绿色建筑验收", "绿建评价", "绿建验收", "绿色建筑")
SURVEY_QUALIFICATION_PATTERN = re.compile(r"测绘.{0,20}乙级|乙级.{0,20}测绘|测绘主管部门.{0,40}乙级")
MONITOR_QUALIFICATION_PATTERN = re.compile(r"工程测量|工程勘察|岩土工程|监测业绩|基坑监测|变形监测|第三方监测|专项监测")
FILTERED_SUFFIX = "-筛选后"
FILTER_CONDITIONS = (1, 2)  # 1=全文检索，2=标题检索；两种都跑，合并去重。
SEARCH_TYPES = (1, 2)  # 1=智能检索，2=精准检索；两种都跑，合并去重。
SCRIPT_DIR = Path(__file__).resolve().parent
DEFAULT_XLSX = SCRIPT_DIR / "2026-乙方宝招标信息统计.xlsx"
DETAIL_SHEET_NAME = "公告详情"
MAIN_SHEET_NAME = "Sheet1"
MAIN_HEADERS = ("序号", "日期", "项目名称", "建设单位", "项目位置", "资质", "报名时间", "投标截止时间", "基本情况", "相关性", "备注")
QUALIFICATION_PLACEHOLDER = "公告资质"
REMARK_PLACEHOLDER = "备注：未完整获取公告正文或未识别到明确资格要求，请人工核对。公告内容"
LIST_WORKERS = 3
DETAIL_WORKERS = 5
MAX_LIST_PAGES = 30
_HTTP_LOCAL = threading.local()


class YfbAuthError(RuntimeError):
    pass


def http_session() -> requests.Session:
    session = getattr(_HTTP_LOCAL, "session", None)
    if session is None:
        session = requests.Session()
        adapter = HTTPAdapter(pool_connections=8, pool_maxsize=8)
        session.mount("https://", adapter)
        session.mount("http://", adapter)
        _HTTP_LOCAL.session = session
    return session


def request_json(path: str, params: dict[str, Any], headers: dict[str, str], timeout: int = 25) -> dict[str, Any]:
    url = f"{BASE}{path}?{urlencode({k: v for k, v in params.items() if v is not None})}"
    last_error: Exception | None = None
    for attempt in range(3):
        try:
            response = http_session().get(url, headers=headers, timeout=timeout)
            if response.status_code == 401:
                raise YfbAuthError("乙方宝接口认证失败，请重新登录")
            response.raise_for_status()
            data = response.json()
            if data.get("code") == 401:
                raise YfbAuthError(data.get("msg") or "乙方宝接口认证失败，请提供登录态")
            if data.get("code") not in (None, 200):
                raise RuntimeError(f"接口返回异常: {data}")
            return data
        except YfbAuthError:
            raise
        except Exception as exc:
            last_error = exc
            if attempt < 2:
                time.sleep(1.5 * (attempt + 1))
    raise RuntimeError(f"请求失败，已重试 3 次: {last_error}")



def request_official_json(path: str, form: dict[str, str]) -> dict[str, Any]:
    headers = {
        "User-Agent": "Mozilla/5.0",
        "Accept": "application/json, text/plain, */*",
        "Origin": "https://cg.95306.cn",
        "Referer": "https://cg.95306.cn/",
        "X-Requested-With": "XMLHttpRequest",
    }
    url = f"https://cg.95306.cn/proxy/portal/elasticSearch{path}"
    body = urlencode(form).encode("utf-8")
    with urlopen(Request(url, data=body, headers=headers), timeout=25) as response:
        data = json.loads(response.read().decode("utf-8", "ignore"))
    if not data.get("success"):
        raise RuntimeError(data.get("msg") or "国铁采购平台查询失败")
    return data


def fetch_official_content(title: str) -> str:
    code_match = re.search(r"20\d{2}(?:-[A-Z0-9]+){5,}", clean_html(title))
    if not code_match:
        return ""
    for attempt in range(5):
        mh_id = uuid.uuid4().hex
        query = {
            "mhId": mh_id,
            "Authorization": "",
            "projBidType": "",
            "bidType": "",
            "noticeType": "000",
            "unitType": "",
            "wzType": "",
            "title": code_match.group(0),
            "inforCode": "",
            "startDate": "",
            "endDate": "",
            "pageNum": "1",
            "projType": "",
            "createPeopUnit": "",
        }
        try:
            result = request_official_json("/queryProcurementNoticeList", query)
            items = result.get("data", {}).get("resultData", {}).get("result", [])
            notice_id = items[0].get("id") if items else ""
            if not notice_id:
                return ""
            detail = request_official_json("/indexView", {
                "noticeId": str(notice_id),
                "mhId": mh_id,
                "Authorization": "",
            })
            content = str(detail.get("data", {}).get("noticeContent", {}).get("notCont") or "")
            if content:
                return content
        except Exception:
            if attempt < 2:
                time.sleep(attempt + 1)
    return ""



def infer_openid(cookie: str) -> str:
    token_match = re.search(r"(?:^|;\s*)Admin-Token=([^;]+)", cookie)
    if not token_match:
        return ""
    try:
        payload = token_match.group(1).split(".")[1]
        payload += "=" * (-len(payload) % 4)
        data = json.loads(base64.urlsafe_b64decode(payload).decode("utf-8"))
        return str(data.get("ei") or "")
    except Exception:
        return ""


def build_headers(args: argparse.Namespace) -> dict[str, str]:
    token = args.token or os.getenv("YFB_TOKEN", "")
    cookie = args.cookie or os.getenv("YFB_COOKIE", "")
    headers = {
        "User-Agent": "Mozilla/5.0",
        "Accept": "application/json, text/plain, */*",
        "Referer": REFERER,
    }
    if token:
        headers["Authorization"] = f"Bearer {token}"
    if cookie:
        headers["Cookie"] = cookie
    return headers


def parse_date(value: Any) -> datetime | None:
    if value is None:
        return None
    if isinstance(value, (int, float)):
        try:
            if value > 10_000_000_000:
                value /= 1000
            return datetime.fromtimestamp(value)
        except Exception:
            return None
    text = str(value).strip()
    if not text:
        return None
    if text in ("今天", "今日"):
        now = datetime.now()
        return datetime(now.year, now.month, now.day)
    if text == "昨天":
        day = datetime.now() - timedelta(days=1)
        return datetime(day.year, day.month, day.day)
    for fmt in ("%Y-%m-%d %H:%M:%S", "%Y/%m/%d %H:%M:%S", "%Y-%m-%d", "%Y/%m/%d"):
        try:
            return datetime.strptime(text[: len(fmt)], fmt)
        except ValueError:
            pass
    m = re.search(r"(20\d{2})\D+(\d{1,2})\D+(\d{1,2})", text)
    if m:
        return datetime(int(m.group(1)), int(m.group(2)), int(m.group(3)))
    return None


def first_value(obj: Any, keys: tuple[str, ...]) -> Any:
    if isinstance(obj, dict):
        for key in keys:
            if obj.get(key):
                return obj[key]
        for value in obj.values():
            found = first_value(value, keys)
            if found:
                return found
    elif isinstance(obj, list):
        for item in obj:
            found = first_value(item, keys)
            if found:
                return found
    return None


def flatten_text(obj: Any) -> str:
    if isinstance(obj, dict):
        return "\n".join(flatten_text(v) for v in obj.values() if v is not None)
    if isinstance(obj, list):
        return "\n".join(flatten_text(v) for v in obj)
    return str(obj)


def clean_html(text: Any) -> str:
    text = "" if text is None else str(text)
    text = re.sub(r"<(?:br|/p|/div|/li|/tr|/h[1-6])\b[^>]*>", "\n", text, flags=re.I)
    text = re.sub(r"<[^>]+>", "", text)
    text = re.sub(r"&nbsp;?", " ", text)
    text = re.sub(r"[ \t]+", " ", text)
    return re.sub(r"\n{3,}", "\n\n", text).strip()


def infer_unit_from_text(title: str, content: str) -> str:
    text = clean_html(content)
    title_text = clean_html(title)
    patterns = (
        r"(?:招标人名称|采购人名称|建设单位名称)\s*[：:]?\s*([\u4e00-\u9fa5A-Za-z0-9（）()·・\-]{2,80})",
        r"(?:采购人|招标人|发包人|建设单位|采购单位|招标单位)\s*(?:为|：|:)\s*([^，,。；;\n]{2,80})",
        r"(?:采购人信息|招标人信息)[\s\S]{0,80}?名称\s*[：:]\s*([^，,。；;\n]{2,80})",
        r"(?:凡对本次(?:采购|招标).*?联系|对本次(?:采购|招标).*?询问)[\s\S]{0,160}?名称\s*[：:]\s*([^，,。；;\n]{2,80})",
    )
    for pattern in patterns:
        match = re.search(pattern, text)
        if match:
            unit = clean_html(match.group(1)).strip(" ：:，,。；;")
            unit = re.split(r"\s+(?:地址|电话|联系方式|联系人)", unit, 1)[0].strip()
            if 2 <= len(unit) <= 80 and not any(bad in unit for bad in ("采购代理", "代理机构", "项目联系人")):
                return unit
    title_patterns = (
        r"^([\u4e00-\u9fa5]{2,50}(?:厅|局|委员会|管理委员会|管理部|办事处|集团|有限公司|研究院|医院|学院|中心|公司))",
        r"^([\u4e00-\u9fa5]{2,50}自然资源局)",
    )
    for pattern in title_patterns:
        match = re.search(pattern, title_text)
        if match:
            unit = match.group(1).strip()
            if "项目" not in unit and len(unit) <= 60:
                return unit
    for marker in ("厅", "自然资源局", "管理委员会城市管理部", "管理委员会建设管理部", "集团"):
        idx = title_text.find(marker)
        if 1 < idx < 40:
            return title_text[: idx + len(marker)]
    return ""


def text_for_keys(obj: Any, keys: tuple[str, ...]) -> str:
    """Find the first usable text field recursively in a detail response."""
    if isinstance(obj, dict):
        for key in keys:
            if obj.get(key):
                return clean_html(obj[key])
        for value in obj.values():
            found = text_for_keys(value, keys)
            if found:
                return found
    elif isinstance(obj, list):
        for value in obj:
            found = text_for_keys(value, keys)
            if found:
                return found
    return ""


def extract_qualification(announcement: str, api_value: str = "") -> str:
    """Extract qualification text across common bidding/government-procurement wording."""
    headings = (
        "投标人资格要求", "供应商资格要求", "申请人的资格要求", "竞标人资格要求",
        "响应人资格要求", "响应供应商资格要求", "报价人资格要求", "报名资格要求",
        "本项目的特定资格要求", "特定资格要求", "申请人应具备的条件", "申请人应满足",
        "合格供应商资格条件", "投标人应具有", "供应商须具备", "投标人须具备",
        "报价人须具备", "响应人须具备", "供应商应具有", "申请人须具备", "资格条件", "资格要求",
    )
    parts = [clean_html(api_value)] if api_value else []
    normalized = clean_html(re.sub(r"\r\n?", "\n", announcement))
    heading_pattern = "|".join(map(re.escape, headings))
    matches = []
    for match in re.finditer(heading_pattern, normalized):
        if match.group() in ("资格要求", "资格条件"):
            prefix = normalized[:match.start()]
            if not re.search(r"(?:^|\n)\s*(?:\d+[.、]|[一二三四五六七八九十]+、|[（(]\d+[)）])?\s*$", prefix):
                continue
        matches.append(match)
    for index, match in enumerate(matches):
        end = matches[index + 1].start() if index + 1 < len(matches) else len(normalized)
        section = normalized[match.start():end]
        stop = re.search(
            r"\n\s*(?:[一二三四五六七八九十]+、|\d+[.、]|[（(]\d+[)）])?\s*"
            r"(?:报名|获取|谈判文件|采购文件|响应文件|投标文件|开标|递交|公告期限|联系方式|其他补充|项目概况)",
            section,
        )
        if stop and stop.start() > 30:
            section = section[:stop.start()]
        section = section.strip()
        if section and section not in parts:
            parts.append(section)
    if not parts:
        patterns = (
            r"[^。；;\n]{0,80}(?:测绘乙级|乙级及以上测绘|测绘主管部门颁发|工程测量|工程勘察|岩土工程|监测业绩|水土保持|绿色建筑评价|绿色建筑验收)[^。；;\n]{0,180}",
            r"[^。；;\n]{0,80}(?:投标人|供应商|申请人|报价人|响应人)(?:应具有|须具有|须具备|应具备)[^。；;\n]{0,220}",
        )
        for pattern in patterns:
            for match in re.finditer(pattern, normalized):
                section = match.group(0).strip(" ：:，,。；;\n")
                if section and section not in parts:
                    parts.append(section)
    return "\n".join(parts)[:12000]


def extract_signup_time(announcement: str, api_value: str = "") -> str:
    parts = [clean_html(api_value)] if api_value else []
    normalized = clean_html(re.sub(r"\r\n?", "\n", announcement))
    date_range = re.compile(
        r"(?:20\d{2}年\s*\d{1,2}月\s*\d{1,2}日|\d{1,2}月\s*\d{1,2}日|20\d{2}[-/]\d{1,2}[-/]\d{1,2})"
        r"[\s\S]{0,80}?(?:至|到|-)\s*"
        r"(?:20\d{2}年\s*)?\d{1,2}(?:月|[-/])\s*\d{1,2}(?:日)?(?:\s*\d{1,2}[：:]\d{2})?"
    )
    for match in date_range.finditer(normalized):
        start = max(0, normalized.rfind("\n", 0, match.start()))
        end = normalized.find("\n", match.end())
        if end == -1:
            end = min(len(normalized), match.end() + 160)
        line = normalized[start:end].strip(" ：:；;，,\n")
        line = re.split(r"(?:获取方式|报名所需|5\.|五、|并将|将以下|邮件|电话通知)", line, 1)[0].strip(" ，,。；;")
        line = re.sub(r"\s+", "", line)
        if any(key in line for key in ("获取", "报名", "发售", "领取", "下载", "请于", "凡有意", "文件")):
            if line and line not in parts:
                parts.append(line)
    return "\n".join(parts)[:1200]


def extract_bid_deadline(announcement: str) -> str:
    normalized = clean_html(re.sub(r"\r\n?", "\n", announcement))
    patterns = (
        r"(?:投标|响应|竞标)文件(?:递交|提交|上传)?(?:的)?(?:截止时间|递交截止时间)[^。；;\n]{0,120}",
        r"(?:投标|响应|竞标)(?:截止时间|截止日期)[^。；;\n]{0,120}",
        r"递交截止时间[^。；;\n]{0,120}",
        r"截止时间(?:为|：|:)?[^。；;\n]{0,120}",
    )
    date_time = re.compile(
        r"20\d{2}年\s*\d{1,2}月\s*\d{1,2}日(?:\s*(?:上午|下午)?\s*\d{1,2}(?:[：:]\d{2}|时\d{2}分?))?"
        r"|20\d{2}[-/]\d{1,2}[-/]\d{1,2}(?:\s*(?:上午|下午)?\s*\d{1,2}[：:]\d{2})?"
    )
    for pattern in patterns:
        for match in re.finditer(pattern, normalized):
            text = re.sub(r"\s+", "", match.group(0)).strip(" ：:，,。；;")
            dt = date_time.search(text)
            if not dt:
                continue
            label = "投标截止时间"
            if "响应" in text:
                label = "响应截止时间"
            elif "竞标" in text:
                label = "竞标截止时间"
            elif "递交" in text:
                label = "递交截止时间"
            return f"{label}：{dt.group(0)}"
    return ""


PROTECTED_TITLE_TERMS = (
    "水土保持", "测绘服务", "测绘项目", "国土测绘", "基础测绘", "地形图", "规划核实",
    "房产实测", "国土变更调查", "用地预审", "确权登记", "土地复垦", "竣工测量",
    "工程测量", "多测合一", "变形监测", "基坑监测", "第三方监测", "专项监测",
    "水土保持监测", "监测、验收", "水土保持验收", "水土保持方案", "遥感监管", "遥感专项监测",
    "绿色建筑评价", "绿色建筑验收",
)

NOISE_TITLE_TERMS = (
    "闲置车位", "车位使用权", "房间招标", "宾馆", "招租", "废旧资产", "机械密封",
    "水电户表", "金属栏杆", "格栅", "爬梯", "中央空调室内机", "加工件",
    "UPS系统采购", "电子反拍", "住宅、储藏室", "山沙一宗", "设备车间设置职工舒缓室",
    "充电站建设项目设计服务", "校园校舍建筑设施安全", "供水管网检测项目", "地基检测项目",
    "测绘仪器无人机采购", "无人机招标公告", "设备健康监测感知层设备与材料采购",
    "交易公告", "转让122套住宅", "联勤宾馆", "联勤宾馆迎宾楼", "双口峪村山沙",
    "莱芜基地用机械密封", "华电国际电力股份有限公司莱城发电厂", "设备租赁",
    "全站仪设备租赁", "监控", "摄像头", "观察孔", "维修件", "仪器维修", "设备采购",
    "过滤器", "水泵", "传动链", "钢丝", "预埋铁座", "模具", "塑料预埋件",
    "商品混凝土", "钢模板", "箱梁模板", "砂石料棚", "钢筋加工棚", "沥青", "河砂",
    "机制砂", "碎石", "铝合金门窗", "锚杆", "施工围挡", "钢丝网护栏", "防护网",
    "管路配件", "抗裂剂", "防火密封胶", "絮凝剂", "战略采购", "地名编制",
    "地图编制", "政务工作用图", "电子地图",
)

STRICT_NOISE_TITLE_TERMS = (
    "测绘仪器无人机采购", "UPS系统采购", "设备健康监测感知层设备与材料采购",
)


def normalize_title_for_dedupe(title: str) -> str:
    text = clean_html(title)
    text = re.sub(r"[（(]原标题[:：].*?[）)]", "", text)
    text = re.sub(r"第\d+次延期", "", text)
    text = re.sub(r"第一次变更公告|第一次更正公示|变更公告|更正公告|二次招标|二次|--.*$", "", text)
    return re.sub(r"\s+", "", text)


def is_noise_title(title: str) -> bool:
    title_text = clean_html(title)
    if any(term in title_text for term in STRICT_NOISE_TITLE_TERMS):
        return True
    if any(term in title_text for term in PROTECTED_TITLE_TERMS):
        return False
    return any(term in title_text for term in NOISE_TITLE_TERMS)

def relevance_level(keyword: str, title: str, text: str) -> str:
    title_text = clean_html(title)
    haystack = clean_html("\n".join(x for x in (title, text) if x))
    if not haystack or is_noise_title(title_text):
        return ""
    if keyword == "水土保持":
        return "明确相关" if any(term in title_text for term in ("水土保持", "水保")) else "疑似相关" if any(term in haystack for term in ("水土保持", "水保")) else ""
    if keyword in ("绿色建筑评价", "绿色建筑验收"):
        green_terms = ("绿色建筑评价", "绿色建筑验收", "绿建评价", "绿建验收")
        return "明确相关" if any(term in title_text for term in green_terms) else "疑似相关" if any(term in haystack for term in green_terms) else ""
    if keyword == "监测":
        title_terms = ("监测", "变形监测", "基坑监测", "沉降观测", "第三方监测", "专项监测")
        body_terms = title_terms + ("监测服务", "监测工作", "监测项目", "深基坑", "基坑支护")
        if any(term in title_text for term in title_terms):
            return "明确相关"
        return "疑似相关" if any(term in haystack for term in body_terms) else ""
    if keyword == "测绘":
        title_terms = ("测绘", "地形图", "规划核实", "房产实测", "国土变更调查", "不动产")
        body_terms = title_terms + ("测绘服务", "测绘成果", "测绘项目")
        if any(term in title_text for term in title_terms):
            return "明确相关"
        return "疑似相关" if any(term in haystack for term in body_terms) else ""
    if keyword == "测量":
        title_terms = ("测量", "工程测量", "竣工测量", "房产实测", "规划核实", "地形图")
        body_terms = title_terms + ("测量服务", "测量工作", "测量项目")
        if any(term in title_text for term in title_terms):
            return "明确相关"
        return "疑似相关" if any(term in haystack for term in body_terms) else ""
    return "明确相关" if keyword in title_text else "疑似相关" if keyword in haystack else ""

def is_relevant_keyword_match(keyword: str, title: str, text: str) -> bool:
    return bool(relevance_level(keyword, title, text))


def is_relevant_monitoring_title(title: str, keyword: str = "", blob: str = "") -> bool:
    return is_relevant_keyword_match(keyword, title, blob)


def one_month_ago_start(today: datetime | None = None) -> datetime:
    today = today or datetime.now()
    month = today.month - 1
    year = today.year
    if month == 0:
        month = 12
        year -= 1
    day = min(today.day, 28 if month == 2 else 30 if month in (4, 6, 9, 11) else 31)
    return datetime(year, month, day)


def list_combo_label(combo: tuple[int, str, int, int]) -> str:
    combo_no, keyword, filter_condition, search_type = combo
    filter_name = "全文检索" if filter_condition == 1 else "标题检索"
    search_name = "智能检索" if search_type == 1 else "精准检索"
    return f"组合 {combo_no}：{keyword} / {filter_name} / {search_name}"


def fetch_list_combo(
    combo: tuple[int, str, int, int],
    headers: dict[str, str],
    openid: str,
    cutoff: datetime,
    time_option: int,
) -> tuple[int, list[dict[str, Any]]]:
    combo_no, keyword, filter_condition, search_type = combo
    label = list_combo_label(combo)
    combo_rows: list[dict[str, Any]] = []
    local_seen: set[str] = set()
    local_seen_titles: set[str] = set()
    for page in range(1, MAX_LIST_PAGES + 1):
        params = {
            "pageSize": 100,
            "pageNum": page,
            "pageFrom": "zhaobiao",
            "keyword": keyword,
            "areaIds": AREA_IDS,
            "filterCondition": filter_condition,
            "searchType": search_type,
            # 长范围用“近三个月”(5)减少无关历史页；短范围用“全部”(4)，再由脚本本地过滤。
            "timeOption": time_option,
            "viewMonitor": "false",
            "defTimeFlag": 0,
            "openid": openid or None,
        }
        try:
            data = request_json("/subZhaobiao/queryZBInfo", params, headers).get("data") or {}
        except YfbAuthError:
            raise
        except Exception as exc:
            raise RuntimeError(f"{label} 第 {page} 页请求失败：{exc}") from exc
        items = data.get("resultList") or data.get("realTimeList") or data.get("resultSet") or []
        if not items:
            print(f"[列表] {label} 第 {page} 页：无结果，组合完成", flush=True)
            break
        old_count = 0
        added_before_page = len(combo_rows)
        for item in items:
            dt = parse_date(first_value(item, ("publishDate", "releaseDate", "createTime", "addTime", "updateDate")))
            area = str(first_value(item, ("areaName", "area", "areaText")) or "")
            blob = flatten_text(item)
            title = clean_html(first_value(item, ("title", "projectName", "name")))
            title_key = normalize_title_for_dedupe(title)
            content_id = str(item.get("contentId") or item.get("id") or title_key or title)
            if dt and dt < cutoff:
                old_count += 1
                continue
            if content_id in local_seen or (title_key and title_key in local_seen_titles):
                continue
            if "济南" not in area and "莱芜" not in area and "济南" not in blob and "莱芜" not in blob:
                continue
            level = relevance_level(keyword, title, blob) or "待终筛"
            candidate = dict(item)
            candidate["_searchKeyword"] = keyword
            candidate["_filterCondition"] = filter_condition
            candidate["_searchType"] = search_type
            candidate["_relevanceLevel"] = level
            combo_rows.append(candidate)
            local_seen.add(content_id)
            if title_key:
                local_seen_titles.add(title_key)
        added_this_page = len(combo_rows) - added_before_page
        print(
            f"[列表] {label} 第 {page} 页：返回 {len(items)} 条，本组合新增 {added_this_page} 条",
            flush=True,
        )
        if old_count == len(items):
            print(f"[列表] {label} 第 {page} 页：结果已早于时间范围，组合完成", flush=True)
            break
    return combo_no, combo_rows


def merge_list_results(results: dict[int, list[dict[str, Any]]]) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    seen: set[str] = set()
    seen_titles: set[str] = set()
    for combo_no in sorted(results):
        for item in results[combo_no]:
            title = clean_html(first_value(item, ("title", "projectName", "name")))
            title_key = normalize_title_for_dedupe(title)
            content_id = str(item.get("contentId") or item.get("id") or title_key or title)
            if content_id in seen or (title_key and title_key in seen_titles):
                continue
            rows.append(item)
            seen.add(content_id)
            if title_key:
                seen_titles.add(title_key)
    return rows


def fetch_list(headers: dict[str, str], openid: str, days: int | None) -> list[dict[str, Any]]:
    today = datetime.now()
    cutoff = datetime(today.year, today.month, today.day) - timedelta(days=days) if days else one_month_ago_start(today)
    time_option = 5 if days and days > 30 else 4
    combos: list[tuple[int, str, int, int]] = []
    combo_no = 0
    for keyword in SEARCH_KEYWORDS:
        for filter_condition in FILTER_CONDITIONS:
            for search_type in SEARCH_TYPES:
                combo_no += 1
                combos.append((combo_no, keyword, filter_condition, search_type))

    total_combos = len(combos)
    print(
        f"[列表] 开始获取列表：关键词 {len(SEARCH_KEYWORDS)} 个，检索组合 {total_combos} 组，"
        f"每组最多 {MAX_LIST_PAGES} 页，列表并发 {LIST_WORKERS} 路",
        flush=True,
    )
    results: dict[int, list[dict[str, Any]]] = {}
    failed_combos: list[tuple[int, str, int, int]] = []
    with ThreadPoolExecutor(max_workers=LIST_WORKERS, thread_name_prefix="yfb-list") as executor:
        future_map = {
            executor.submit(fetch_list_combo, combo, headers, openid, cutoff, time_option): combo
            for combo in combos
        }
        for completed, future in enumerate(as_completed(future_map), start=1):
            combo = future_map[future]
            try:
                completed_no, combo_rows = future.result()
                results[completed_no] = combo_rows
                print(
                    f"[列表] 已完成 {completed}/{total_combos}：{list_combo_label(combo)}，候选 {len(combo_rows)} 条",
                    flush=True,
                )
            except YfbAuthError:
                raise
            except Exception as exc:
                failed_combos.append(combo)
                print(
                    f"[列表] 并发抓取失败，加入串行补抓：{list_combo_label(combo)}；{exc}",
                    file=sys.stderr,
                    flush=True,
                )

    if failed_combos:
        print(f"[列表] 开始串行补抓失败组合，共 {len(failed_combos)} 组", flush=True)
    for retry_index, combo in enumerate(failed_combos, start=1):
        try:
            completed_no, combo_rows = fetch_list_combo(combo, headers, openid, cutoff, time_option)
            results[completed_no] = combo_rows
            print(
                f"[列表] 补抓成功 {retry_index}/{len(failed_combos)}：{list_combo_label(combo)}",
                flush=True,
            )
        except YfbAuthError:
            raise
        except Exception as exc:
            raise RuntimeError(
                f"列表检索未完整完成，串行补抓仍失败：{list_combo_label(combo)}；{exc}"
            ) from exc

    if len(results) != total_combos:
        missing = [list_combo_label(combo) for combo in combos if combo[0] not in results]
        raise RuntimeError(f"列表检索结果不完整，缺少组合：{'；'.join(missing)}")

    rows = merge_list_results(results)
    print(f"[列表] 全部 {total_combos} 个组合完整结束：去重后候选 {len(rows)} 条，开始排序", flush=True)
    rows.sort(
        key=lambda item: parse_date(
            first_value(item, ("publishDate", "releaseDate", "createTime", "addTime", "updateDate"))
        ) or datetime.min,
        reverse=True,
    )
    return rows
def fetch_detail(
    item: dict[str, Any],
    headers: dict[str, str],
    openid: str,
    strict: bool = False,
) -> dict[str, Any]:
    content_id = item.get("contentId") or item.get("id")
    area_id = item.get("areaId") or item.get("area")
    if not content_id:
        return {}
    params = {"contentId": content_id, "areaId": area_id, "pageFrom": "search", "openid": openid or None}
    try:
        detail = request_json("/subZhaobiao/zbDetail", params, headers, timeout=8).get("data") or {}
    except YfbAuthError:
        raise
    except Exception:
        if strict:
            raise
        detail = {}
    detail_content = text_for_keys(detail, ("content", "noticeContent", "htmlContent", "detailContent", "text"))
    item_content = text_for_keys(item, ("content", "summary", "noticeContent", "htmlContent", "text"))
    qualification = text_for_keys(detail, ("qualification", "aptitude", "qualificationRequirement"))
    combined_content = "\n".join(x for x in (detail_content, item_content) if x)
    if not extract_qualification(combined_content, qualification):
        official_content = fetch_official_content(str(item.get("title") or ""))
        if official_content:
            detail["officialContent"] = official_content
    return detail


def fetch_details_with_retry(
    pending_details: list[tuple[int, dict[str, Any], str]],
    headers: dict[str, str],
    openid: str,
) -> list[tuple[int, dict[str, Any], str, dict[str, Any]]]:
    if not pending_details:
        return []
    print(f"[详情] 启用 {DETAIL_WORKERS} 路并发，待获取 {len(pending_details)} 条", flush=True)
    detail_results: list[tuple[int, dict[str, Any], str, dict[str, Any]]] = []
    failed_details: list[tuple[int, dict[str, Any], str, Exception]] = []
    with ThreadPoolExecutor(max_workers=DETAIL_WORKERS, thread_name_prefix="yfb-detail") as executor:
        future_map = {
            executor.submit(fetch_detail, item, headers, openid, True): (detail_index, item, list_title)
            for detail_index, item, list_title in pending_details
        }
        for completed, future in enumerate(as_completed(future_map), start=1):
            detail_index, item, list_title = future_map[future]
            try:
                detail = future.result()
                detail_results.append((detail_index, item, list_title, detail))
                status = "完成"
            except YfbAuthError:
                raise
            except Exception as exc:
                failed_details.append((detail_index, item, list_title, exc))
                status = "失败，等待补抓"
            print(
                f"[详情] 已处理 {completed}/{len(pending_details)}（{status}）：{list_title[:80]}",
                flush=True,
            )

    if failed_details:
        print(f"[详情] 开始串行补抓失败项目，共 {len(failed_details)} 条", flush=True)
    for retry_index, (detail_index, item, list_title, first_error) in enumerate(failed_details, start=1):
        try:
            detail = fetch_detail(item, headers, openid, True)
            detail_results.append((detail_index, item, list_title, detail))
            print(
                f"[详情] 补抓成功 {retry_index}/{len(failed_details)}：{list_title[:80]}",
                flush=True,
            )
        except YfbAuthError:
            raise
        except Exception as retry_error:
            item["_detailFetchError"] = f"首次：{first_error}；补抓：{retry_error}"
            detail_results.append((detail_index, item, list_title, {}))
            print(
                f"[详情] 补抓仍失败，保留列表信息并标记人工核对：{list_title[:80]}；{retry_error}",
                file=sys.stderr,
                flush=True,
            )
    detail_results.sort(key=lambda result: result[0])
    return detail_results

def row_from_item(item: dict[str, Any], detail: dict[str, Any]) -> list[Any]:
    merged = {"list": item, "detail": detail}
    dt = parse_date(first_value(merged, ("publishDate", "releaseDate", "createTime", "addTime", "updateDate")))
    title = clean_html(first_value(merged, ("title", "projectName", "name")))
    unit = clean_html(first_value(merged, ("zhaoBiaoUnit", "zhaoBiaoRen", "tenderer", "buyerName", "purchaseUnit")))
    area = clean_html(first_value(merged, ("areaName", "areaText", "area")))
    content = text_for_keys(detail, ("officialContent", "content", "noticeContent", "htmlContent", "detailContent", "text"))
    if not content:
        content = text_for_keys(item, ("content", "summary", "noticeContent", "htmlContent", "text"))
    if not unit:
        unit = infer_unit_from_text(title, content)
    if not unit:
        match = re.match(r"(.{2,40}?有限公司)", title)
        if match:
            unit = match.group(1).strip()
    qualification = text_for_keys(detail, ("qualification", "aptitude", "qualificationRequirement"))
    signup_time = clean_html(first_value(merged, ("signUpTime", "registrationTime", "tenderEndTimeStr", "bidEndTime")))
    qualification_text = extract_qualification(content, qualification)
    time_text = extract_signup_time(content, signup_time)
    deadline_text = extract_bid_deadline(content)
    url = f"https://qiye.qianlima.com/new_qd_yfbsite/#/infoCenter/infoDetail/{item.get('contentId')}/{item.get('areaId')}/zhaobiao"
    return [
        None,
        dt.strftime("%m.%d") if dt else "",
        title,
        unit,
        area,
        qualification_text or "未在公告正文中识别到明确的资格要求",
        time_text,
        deadline_text,
        "",
        item.get("_relevanceLevel") or "明确相关",
        url,
    ]



def ensure_relevance_column(ws: Any) -> None:
    if ws.cell(1, 10).value == "相关性":
        return
    ws.insert_cols(10)
    ws.cell(1, 10, "相关性")
    for row in range(1, ws.max_row + 1):
        src = ws.cell(row, 9)
        dst = ws.cell(row, 10)
        if src.has_style:
            dst._style = copy.copy(src._style)
        if src.number_format:
            dst.number_format = src.number_format
        if src.alignment:
            dst.alignment = copy.copy(src.alignment)


def ensure_main_sheet(wb: Any) -> Any:
    ws = next(
        (
            sheet for sheet in wb.worksheets
            if sheet.title != DETAIL_SHEET_NAME
            and str(sheet.cell(1, 3).value or "").strip() == "项目名称"
        ),
        None,
    )
    if ws is None and MAIN_SHEET_NAME in wb.sheetnames:
        ws = wb[MAIN_SHEET_NAME]
    if ws is None:
        ws = next(
            (
                sheet for sheet in wb.worksheets
                if sheet.title != DETAIL_SHEET_NAME
                and all(sheet.cell(1, col).value in (None, "") for col in range(1, 12))
            ),
            None,
        )
    if ws is None:
        ws = wb.create_sheet(MAIN_SHEET_NAME, 0)
    if str(ws.cell(1, 3).value or "").strip() != "项目名称":
        for col, header in enumerate(MAIN_HEADERS, start=1):
            ws.cell(1, col, header)
            ws.cell(1, col).font = Font(bold=True)
        ws.freeze_panes = "A2"
    wb.active = wb.index(ws)
    return ws


def ensure_detail_sheet(wb: Any) -> Any:
    if DETAIL_SHEET_NAME in wb.sheetnames:
        detail_ws = wb[DETAIL_SHEET_NAME]
    else:
        detail_ws = wb.create_sheet(DETAIL_SHEET_NAME)
        detail_ws.append(["主表序号", "项目名称", "内容类型", "详细内容", "返回主表"])
    headers = ["主表序号", "项目名称", "内容类型", "详细内容", "返回主表"]
    for col, header in enumerate(headers, start=1):
        detail_ws.cell(1, col, header)
        detail_ws.cell(1, col).font = Font(bold=True)
    widths = {1: 12, 2: 70, 3: 16, 4: 120, 5: 18}
    for column, width in widths.items():
        letter = detail_ws.cell(1, column).column_letter
        detail_ws.column_dimensions[letter].width = width
    detail_ws.freeze_panes = "A2"
    trim_empty_tail(detail_ws, 2)
    return detail_ws


def short_remark_label(text: str) -> str:
    text = clean_html(text)
    if "未完整获取公告正文" in text or "未识别到明确资格要求" in text or "公告内容" in text:
        return REMARK_PLACEHOLDER
    if "疑似相关" in text:
        return "备注：疑似相关，需人工核对"
    return "公告备注"


def set_internal_hyperlink(cell: Any, location: str, label: str | None = None) -> None:
    if label is not None:
        cell.value = label
    cell.hyperlink = Hyperlink(ref=cell.coordinate, location=location)
    cell.style = "Hyperlink"


def add_detail_link(ws: Any, detail_ws: Any, row: int, col: int, title: str, kind: str, full_text: str, label: str) -> None:
    full_text = clean_html(full_text)
    if not full_text:
        return
    detail_row = detail_ws.max_row + 1
    main_cell = ws.cell(row, col)
    detail_ws.cell(detail_row, 1, ws.cell(row, 1).value)
    detail_ws.cell(detail_row, 2, title)
    detail_ws.cell(detail_row, 3, kind)
    detail_ws.cell(detail_row, 4, full_text)
    back_cell = detail_ws.cell(detail_row, 5, "返回主表")
    set_internal_hyperlink(back_cell, f"'{ws.title}'!{main_cell.coordinate}")
    for c in range(1, 6):
        cell = detail_ws.cell(detail_row, c)
        alignment = copy.copy(cell.alignment)
        alignment.wrap_text = True
        alignment.vertical = "top"
        cell.alignment = alignment
    detail_ws.row_dimensions[detail_row].height = min(409, max(45, len(full_text) // 80 * 15))
    set_internal_hyperlink(main_cell, f"'{DETAIL_SHEET_NAME}'!D{detail_row}", label)
    alignment = copy.copy(main_cell.alignment)
    alignment.wrap_text = True
    alignment.vertical = "top"
    main_cell.alignment = alignment


def move_long_text_to_detail(ws: Any, detail_ws: Any, row: int) -> None:
    title = str(ws.cell(row, 3).value or "")
    qualification = str(ws.cell(row, 6).value or "")
    remark = str(ws.cell(row, 11).value or "") if ws.max_column >= 11 else ""
    if qualification and qualification != QUALIFICATION_PLACEHOLDER:
        add_detail_link(ws, detail_ws, row, 6, title, "资质", qualification, QUALIFICATION_PLACEHOLDER)
    if remark and not remark.startswith("http") and remark not in ("公告备注", REMARK_PLACEHOLDER, "备注：疑似相关，需人工核对"):
        add_detail_link(ws, detail_ws, row, 11, title, "备注", remark, short_remark_label(remark))


def format_notice_rows(ws: Any, row_numbers: list[int]) -> None:
    widths = {1: 9, 2: 13, 3: 62, 4: 37, 5: 18, 6: 60, 7: 34, 8: 30, 9: 30, 10: 14, 11: 55}
    line_widths = {3: 45, 4: 27, 5: 16, 6: 60, 7: 30, 8: 26, 9: 30, 10: 10, 11: 55}
    for column, width in widths.items():
        letter = ws.cell(1, column).column_letter
        current = ws.column_dimensions[letter].width or 0
        ws.column_dimensions[letter].width = max(current, width)
    for row in row_numbers:
        estimated_lines = 1
        for column in range(1, 12):
            cell = ws.cell(row, column)
            alignment = copy.copy(cell.alignment)
            alignment.wrap_text = True
            alignment.vertical = "top"
            cell.alignment = alignment
            value = str(cell.value or "")
            width = line_widths.get(column, 30)
            estimated_lines = max(
                estimated_lines,
                sum(max(1, (len(line) + width - 1) // width) for line in value.splitlines() or [""]),
            )
        ws.row_dimensions[row].height = min(409, max(30, estimated_lines * 15))


def detail_texts_by_title(detail_ws: Any | None) -> dict[str, str]:
    if detail_ws is None:
        return {}
    result: dict[str, str] = {}
    for row in range(2, detail_ws.max_row + 1):
        title = str(detail_ws.cell(row, 2).value or "").strip()
        if not title:
            continue
        content = str(detail_ws.cell(row, 4).value or "")
        result[title] = (result.get(title, "") + "\n" + content).strip()
    return result


def title_matches_business_keyword(title: str) -> bool:
    title_text = clean_html(title)
    title_text = re.sub(r"\s+", "", title_text)
    return any(term in title_text for term in TITLE_KEEP_TERMS)


def filter_decision(title: str, qualification: str, remark: str, detail_text: str) -> tuple[bool, str]:
    title_text = clean_html(title)
    qualification_text = clean_html(qualification)
    full_text = clean_html("\n".join(x for x in (title, qualification, remark, detail_text) if x))
    full_text = re.sub(r"\s+", "", full_text)
    if title_matches_business_keyword(title_text):
        return True, "标题命中业务关键词"
    if any(term in full_text for term in WATER_GREEN_TERMS):
        return True, "公告命中水土保持/绿色建筑，直接保留"
    survey_terms = ("测绘", "测量", "多测合一", "地形图", "规划核实", "房产实测", "国土变更调查", "不动产测绘", "竣工测量")
    monitor_terms = ("基坑监测", "深基坑监测", "第三方监测", "专项监测", "结构监测", "运营期结构监测", "监测服务", "变形监测", "沉降观测", "沉降监测", "周边环境监测")
    if any(term in full_text for term in survey_terms) and SURVEY_QUALIFICATION_PATTERN.search(full_text):
        return True, "公告命中测绘测量且资质含测绘乙级"
    if any(term in full_text for term in monitor_terms) and MONITOR_QUALIFICATION_PATTERN.search(full_text):
        return True, "公告命中基坑监测且资质含工程测量/工程勘察/岩土工程/监测业绩"
    return False, "未命中筛选规则"


def rebuild_internal_links(wb: Any) -> None:
    ws = ensure_main_sheet(wb)
    detail_ws = wb[DETAIL_SHEET_NAME] if DETAIL_SHEET_NAME in wb.sheetnames else None
    for row in range(2, ws.max_row + 1):
        ws.cell(row, 1).value = row - 1
    if detail_ws is None:
        return
    main_rows = {
        str(ws.cell(row, 3).value or "").strip(): row
        for row in range(2, ws.max_row + 1)
        if str(ws.cell(row, 3).value or "").strip()
    }
    detail_rows: dict[tuple[str, str], int] = {}
    for row in range(2, detail_ws.max_row + 1):
        title = str(detail_ws.cell(row, 2).value or "").strip()
        kind = str(detail_ws.cell(row, 3).value or "").strip()
        if title and kind in ("资质", "备注"):
            detail_rows.setdefault((title, kind), row)

    for row in range(2, ws.max_row + 1):
        title = str(ws.cell(row, 3).value or "").strip()
        qualification_cell = ws.cell(row, 6)
        remark_cell = ws.cell(row, 11)
        qualification_cell.hyperlink = None
        remark_cell.hyperlink = None
        qualification_row = detail_rows.get((title, "资质"))
        if qualification_row:
            set_internal_hyperlink(
                qualification_cell,
                f"'{DETAIL_SHEET_NAME}'!D{qualification_row}",
                QUALIFICATION_PLACEHOLDER,
            )
        remark_row = detail_rows.get((title, "备注"))
        if remark_row:
            remark_text = str(detail_ws.cell(remark_row, 4).value or "")
            set_internal_hyperlink(
                remark_cell,
                f"'{DETAIL_SHEET_NAME}'!D{remark_row}",
                short_remark_label(remark_text),
            )

    for row in range(2, detail_ws.max_row + 1):
        title = str(detail_ws.cell(row, 2).value or "").strip()
        kind = str(detail_ws.cell(row, 3).value or "").strip()
        main_row = main_rows.get(title)
        back_cell = detail_ws.cell(row, 5)
        back_cell.hyperlink = None
        if main_row:
            detail_ws.cell(row, 1).value = ws.cell(main_row, 1).value
            main_col = 11 if kind == "备注" else 6
            set_internal_hyperlink(
                back_cell,
                f"'{ws.title}'!{ws.cell(main_row, main_col).coordinate}",
                "返回主表",
            )


def trim_empty_tail(ws: Any, title_col: int) -> None:
    last = 1
    for row in range(2, ws.max_row + 1):
        if str(ws.cell(row, title_col).value or "").strip():
            last = row
    for merged_range in list(ws.merged_cells.ranges):
        if merged_range.min_row > last:
            ws.unmerge_cells(str(merged_range))
    if ws.max_row > last:
        ws.delete_rows(last + 1, ws.max_row - last)


def filtered_output_path(path: Path) -> Path:
    return path.with_name(f"{path.stem}{FILTERED_SUFFIX}{path.suffix}")


def worksheet_date_sort_key(value: Any) -> tuple[int, int, str]:
    text = str(value or "").strip()
    match = re.fullmatch(r"(\d{1,2})[.\-/](\d{1,2})", text)
    if match:
        return int(match.group(1)), int(match.group(2)), text
    parsed = parse_date(value)
    if parsed:
        return parsed.month, parsed.day, text
    return 99, 99, text


def expand_date_merges(ws: Any) -> None:
    date_merges = [
        merged
        for merged in list(ws.merged_cells.ranges)
        if merged.min_col == 2 and merged.max_col == 2 and merged.min_row >= 2
    ]
    for merged in date_merges:
        date_value = ws.cell(merged.min_row, 2).value
        min_row, max_row = merged.min_row, merged.max_row
        ws.unmerge_cells(str(merged))
        for row in range(min_row, max_row + 1):
            ws.cell(row, 2).value = date_value


def merge_equal_date_cells(ws: Any) -> None:
    if ws.max_row < 2:
        return
    group_start = 2
    current_date = ws.cell(2, 2).value
    for row in range(3, ws.max_row + 2):
        date_value = ws.cell(row, 2).value if row <= ws.max_row else None
        if date_value == current_date:
            continue
        if current_date not in (None, "") and row - group_start > 1:
            ws.merge_cells(start_row=group_start, start_column=2, end_row=row - 1, end_column=2)
        group_start = row
        current_date = date_value


def sort_main_sheet_by_date(ws: Any) -> None:
    expand_date_merges(ws)
    data = [
        [ws.cell(row, col).value for col in range(1, ws.max_column + 1)]
        for row in range(2, ws.max_row + 1)
        if str(ws.cell(row, 3).value or "").strip()
    ]
    data.sort(key=lambda values: worksheet_date_sort_key(values[1] if len(values) > 1 else ""))
    for target_row, values in enumerate(data, start=2):
        values[0] = target_row - 1
        for col, value in enumerate(values, start=1):
            ws.cell(target_row, col).value = value
    trim_empty_tail(ws, 3)
    merge_equal_date_cells(ws)

def create_filtered_workbook(path: Path) -> Path:
    output = filtered_output_path(path)
    if output.exists():
        try:
            output.unlink()
        except PermissionError as exc:
            raise PermissionError(f"无法覆盖筛选后文件，请先关闭：{output}") from exc
    shutil.copy2(path, output)
    wb = load_workbook(output)
    ws = ensure_main_sheet(wb)
    detail_ws = wb[DETAIL_SHEET_NAME] if DETAIL_SHEET_NAME in wb.sheetnames else None
    details = detail_texts_by_title(detail_ws)
    expand_date_merges(ws)
    deleted_sheet_name = "筛选删除记录"
    if deleted_sheet_name in wb.sheetnames:
        del wb[deleted_sheet_name]
    deleted_ws = wb.create_sheet(deleted_sheet_name)
    headers = [ws.cell(1, col).value for col in range(1, ws.max_column + 1)]
    deleted_ws.append(headers + ["筛出原因"])
    remove_titles: set[str] = set()
    removed = 0
    kept = 0
    for row in range(2, ws.max_row + 1):
        title = str(ws.cell(row, 3).value or "").strip()
        if not title:
            continue
        qualification = str(ws.cell(row, 6).value or "")
        remark = str(ws.cell(row, 11).value or "") if ws.max_column >= 11 else ""
        keep, reason = filter_decision(title, qualification, remark, details.get(title, ""))
        if keep:
            kept += 1
            if ws.max_column >= 10:
                ws.cell(row, 10).value = reason
        else:
            remove_titles.add(title)
            deleted_ws.append([ws.cell(row, col).value for col in range(1, ws.max_column + 1)] + [reason])
    for row in range(ws.max_row, 1, -1):
        title = str(ws.cell(row, 3).value or "").strip()
        if not title or title in remove_titles:
            ws.delete_rows(row, 1)
            if title:
                removed += 1
    if detail_ws is not None:
        for row in range(detail_ws.max_row, 1, -1):
            title = str(detail_ws.cell(row, 2).value or "").strip()
            if not title or title in remove_titles:
                detail_ws.delete_rows(row, 1)
        trim_empty_tail(detail_ws, 2)
    trim_empty_tail(ws, 3)
    trim_empty_tail(deleted_ws, 3)
    merge_equal_date_cells(ws)
    for col in range(1, deleted_ws.max_column + 1):
        deleted_ws.column_dimensions[deleted_ws.cell(1, col).column_letter].width = min(60, max(12, len(str(deleted_ws.cell(1, col).value or "")) + 4))
    rebuild_internal_links(wb)
    wb.save(output)
    print(f"已生成筛选后文件：{output}（保留 {kept} 条，删除 {removed} 条；删除记录见《{deleted_sheet_name}》）")
    return output


def append_to_workbook(path: Path, rows: list[list[Any]], dry_run: bool) -> None:
    if not rows:
        print("未抓到符合条件的新数据，原始 Excel 未修改。")
        if not dry_run and path.exists():
            create_filtered_workbook(path)
        return
    wb = load_workbook(path)
    ws = ensure_main_sheet(wb)
    ensure_relevance_column(ws)
    detail_ws = ensure_detail_sheet(wb)
    # Expand date merges before deleting trailing rows so a merge cannot be
    # partially removed and leave openpyxl with stale merged-cell metadata.
    expand_date_merges(ws)
    while ws.max_row > 2 and all(ws.cell(ws.max_row, c).value in (None, "") for c in range(1, ws.max_column + 1)):
        ws.delete_rows(ws.max_row)
    existing_titles = {
        normalize_title_for_dedupe(str(ws.cell(r, 3).value or ""))
        for r in range(2, ws.max_row + 1)
        if str(ws.cell(r, 3).value or "").strip()
    }
    filtered_rows = []
    seen_new_titles: set[str] = set()
    for row in rows:
        title_key = normalize_title_for_dedupe(str(row[2] or ""))
        if title_key and (title_key in existing_titles or title_key in seen_new_titles):
            continue
        filtered_rows.append(row)
        if title_key:
            seen_new_titles.add(title_key)
    rows = filtered_rows
    if not rows:
        print("抓到的数据均已存在，原始 Excel 未修改。")
        if not dry_run and path.exists():
            create_filtered_workbook(path)
        return
    rows.sort(key=lambda row: worksheet_date_sort_key(row[1] if len(row) > 1 else ""))
    template_row = ws.max_row == 2 and all(
        ws.cell(2, c).value in (None, "") for c in range(1, ws.max_column + 1)
    )
    start = 2 if template_row else ws.max_row + 1
    last_no = max([ws.cell(r, 1).value for r in range(2, ws.max_row + 1) if isinstance(ws.cell(r, 1).value, int)] or [0])
    style_row = 2 if template_row else ws.max_row
    for offset, row in enumerate(rows, start=0):
        target = start + offset
        row[0] = last_no + offset + 1
        for col, value in enumerate(row, start=1):
            cell = ws.cell(target, col, value)
            src = ws.cell(style_row, col)
            if src.has_style:
                cell._style = copy.copy(src._style)
            if src.number_format:
                cell.number_format = src.number_format
            if src.alignment:
                cell.alignment = copy.copy(src.alignment)
        move_long_text_to_detail(ws, detail_ws, target)
    format_notice_rows(ws, list(range(start, start + len(rows))))
    sort_main_sheet_by_date(ws)
    format_notice_rows(ws, list(range(2, ws.max_row + 1)))
    rebuild_internal_links(wb)
    if dry_run:
        print(f"演练模式：将追加 {len(rows)} 条，Excel 未修改。")
        return
    backup = path.with_suffix(f".backup-{datetime.now():%Y%m%d-%H%M%S}.xlsx")
    shutil.copy2(path, backup)
    wb.save(path)
    print(f"已追加 {len(rows)} 条，备份文件：{backup}")
    create_filtered_workbook(path)


def main() -> int:
    parser = argparse.ArgumentParser(description="搜索乙方宝近 N 天山东济南/莱芜监测、水土保持、测绘、测量、绿色建筑招标并追加到 Excel，并生成筛选后文件")
    parser.add_argument("--xlsx", type=Path, default=DEFAULT_XLSX)
    parser.add_argument("--days", type=int, default=7, help="按最近 N 天筛选，默认 7 天")
    parser.add_argument("--token", default="")
    parser.add_argument("--cookie", default="")
    parser.add_argument("--openid", default=os.getenv("YFB_OPENID", ""))
    parser.add_argument("--dry-run", action="store_true")
    args = parser.parse_args()

    headers = build_headers(args)
    openid = args.openid or infer_openid(args.cookie or os.getenv("YFB_COOKIE", ""))
    try:
        existing_titles: set[str] = set()
        if args.xlsx.exists():
            wb_existing = load_workbook(args.xlsx, read_only=True, data_only=True)
            ws_existing = wb_existing.active
            existing_titles = {
                normalize_title_for_dedupe(str(ws_existing.cell(r, 3).value or ""))
                for r in range(2, ws_existing.max_row + 1)
                if str(ws_existing.cell(r, 3).value or "").strip()
            }
        items = fetch_list(headers, openid, args.days)
        print(f"[详情] 列表候选 {len(items)} 条，开始获取详情和解析字段", flush=True)
        rows = []
        skipped_existing = 0
        pending_details = []
        for detail_index, item in enumerate(items, start=1):
            list_title = clean_html(first_value(item, ("title", "projectName", "name")))
            if list_title and normalize_title_for_dedupe(list_title) in existing_titles:
                skipped_existing += 1
                if skipped_existing == 1 or skipped_existing % 10 == 0:
                    print(f"[详情] 已跳过已有项目 {skipped_existing} 条", flush=True)
                continue
            pending_details.append((detail_index, item, list_title))

        detail_results = fetch_details_with_retry(pending_details, headers, openid)
        for detail_index, item, list_title, detail in detail_results:
            title = clean_html(first_value({"list": item, "detail": detail}, ("title", "projectName", "name")))
            content = text_for_keys(detail, ("officialContent", "content", "noticeContent", "htmlContent", "detailContent", "text"))
            if not content:
                content = text_for_keys(item, ("content", "summary", "noticeContent", "htmlContent", "text"))
            keyword = str(item.get("_searchKeyword") or "")
            level = relevance_level(keyword, title, content) if keyword else str(item.get("_relevanceLevel") or "")
            if not level:
                level = "待终筛"
            item["_relevanceLevel"] = level
            row = row_from_item(item, detail)
            remark_parts = []
            if item.get("_detailFetchError"):
                remark_parts.append("备注：乙方宝详情接口并发抓取和串行补抓均失败，当前仅保留列表信息，请点击原公告人工核对。")
            if level == "疑似相关":
                remark_parts.append("备注：疑似相关，需人工核对：关键词未在标题中明确命中，但公告正文或列表摘要中出现相关内容。")
            elif level == "待终筛":
                remark_parts.append("备注：宽口径写入，未在列表或详情中命中明确业务词，交由筛选后表规则判断。")
            if row[5] == "未在公告正文中识别到明确的资格要求":
                print(f"写入但标记为需人工核对：{row[2]}", file=sys.stderr)
                notice_text = clean_html(content) or clean_html(flatten_text(item))
                remark_parts.append("备注：未完整获取公告正文或未识别到明确资格要求，请人工核对。\n公告内容：\n" + notice_text[:10000])
            if remark_parts:
                row[10] = "\n".join(remark_parts) + "\n链接：" + str(row[10] or "")
            rows.append(row)
        print(f"[写入] 详情解析完成：准备写入 {len(rows)} 条新数据，跳过已有 {skipped_existing} 条", flush=True)
        append_to_workbook(args.xlsx, rows, args.dry_run)
    except YfbAuthError as exc:
        print(f"认证失败：{exc}", file=sys.stderr)
        print("请登录乙方宝后提供 YFB_TOKEN，或设置 YFB_COOKIE / YFB_OPENID 再运行。", file=sys.stderr)
        return 2
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
